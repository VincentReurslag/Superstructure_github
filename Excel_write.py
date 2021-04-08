# -*- coding: utf-8 -*-
"""
Created on Mon Apr  5 12:46:43 2021

@author: Vincent Reurslag
"""

from openpyxl import Workbook
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

def excel_write(Superstructure, output_file):
    filename = output_file

    workbook = Workbook()
    sheet = workbook.active
    
    #Make numpy arrays for even spacing in excel sheet to write equipment data
    start_col = 5
    step_col = 5
    col = np.arange(start_col, len(Superstructure.j)*step_col+start_col, step_col)
    start_row = 5
    step_row = 8
    row = np.arange(start_row, len(Superstructure.k)*step_row+start_row, step_row)
    
    #Size of the data square for equpiment types that is colored
    h_size = 3
    v_size = 6
    
    #Defining colors so I can use them easier later on
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type = "solid")
    grey = PatternFill(start_color = "D3D3D3", end_color = "D3D3D3",  fill_type = "solid")
    blue = PatternFill(start_color = "00A0FF", end_color = "00A0FF",  fill_type = "solid")
    
    #Print objective function to the excel sheet
    cell = get_column_letter(start_col) + str(1)
    cell1 = get_column_letter(start_col+h_size) + str(1)
    sheet.merge_cells(cell + ':' + cell1)
    sheet[cell] = 'Total equipment cost: %s' % Superstructure.TEC
    sheet[cell].fill = blue
    
    
    #Write results to Excel
    for j in range(len(Superstructure.j)):
        for k in range(len(Superstructure.k)):
            selected = Superstructure.Logic_data.iloc[k].iloc[j]
            equipment = Superstructure.equipment_names.iloc[k].iloc[j]
            
            
            #Define the 4 corner cells to draw a box in excel
            cell = get_column_letter(col[j]) + str(row[k])
            cell1 = get_column_letter(col[j] + h_size) + str(row[k])
            cell2 = get_column_letter(col[j]) + str(row[k] + v_size)
            cell3 = get_column_letter(col[j] + h_size) + str(row[k] + v_size)
            
            #Merge and center the equipment name cell
            sheet.merge_cells(cell + ':' + cell1)
            sheet[cell].alignment = Alignment(horizontal = 'center')
            sheet[cell] = equipment
            
            #Write for every j,k the component names and corresponding flow data
            sheet[get_column_letter(col[j]) + str(row[k]+1)] = 'Species i'
            sheet[get_column_letter(col[j]+1) + str(row[k]+1)] = 'Flow [kg/h]'
        
            for i in range(len(Superstructure.component_names)):
                sheet[get_column_letter(col[j]) + str(row[k]+i+2)] = Superstructure.component_names.iloc[i]
                sheet[get_column_letter(col[j] + 1) + str(row[k]+i+2)] = Superstructure.Flow_data[j+1][k+1][i+1]
                
            
            #If an equipment is selected mark it green and grey
            if selected == 1:
                sheet[cell].fill = green
                
                for r in range(v_size):
                    for c in range(h_size + 1):
                        cell = get_column_letter(col[j] + c) + str(row[k]+r+1)
                        sheet[cell].fill = grey
                        
    
    workbook.save(filename = filename)