# -*- coding: utf-8 -*-
"""
Created on Sat Apr  3 15:21:03 2021
Run this script to execute te entire program

@author: Vincent Reurslag


"""

import pandas as pd
from Superstructure_class import Superstructure
from Superstructure_model import Superstructure_model
from Excel_write import excel_write
from pyomo.core import ConcreteModel, Set, NonNegativeReals, Var, value


#xlsx_file is the input file with the data, output_file will be the result file
xlsx_file = 'SS_dataV_AN.xlsx'
output_file = 'ResultsV3.xlsx'

a = [1,2,3,4,5,6,7,8]
j = [1,2,3,4]
k = [1,2,3]


Superstructure = Superstructure(xlsx_file,a,j,k)
Superstructure.get_4index(xlsx_file,['SF a,j,k,i', 'Q a,j,k,i','Tau a,j,k,u'])
Superstructure.get_3index(xlsx_file, ['EC a,j,k','Temperature a,j,k','ReferenceCost a,j,k','ReferenceSize a,j,k','ReferenceIndex a,j,k','SizingFactor a,j,k'])
Superstructure.get_1index(xlsx_file, ['Flow0 i', 'CP i', 'CompCost i', 'SpecificCostU u'])

##############################################################
model = Superstructure_model(Superstructure)
Superstructure.get_results(model)


from openpyxl import Workbook
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment




filename = output_file

workbook = Workbook()
sheet = workbook.active

#Make numpy arrays for even spacing in excel sheet to write equipment data
start_col = 1
step_col = 5
col = np.arange(start_col, len(Superstructure.a)*step_col+start_col, step_col)
start_row = 5
step_row = 8
row = np.arange(start_row, len(Superstructure.k)*step_row+start_row, step_row)

h_spacer = 10

#Size of the data square for equpiment types that is colored
h_size = 3
v_size = 8

#Defining colors so I can use them easier later on
green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type = "solid")
grey = PatternFill(start_color = "D3D3D3", end_color = "D3D3D3",  fill_type = "solid")
blue = PatternFill(start_color = "00A0FF", end_color = "00A0FF",  fill_type = "solid")

#Print objective function to the excel sheet
cell = get_column_letter(start_col) + str(start_row + h_spacer)
cell1 = get_column_letter(start_col+5) + str(start_row + h_spacer)
sheet.merge_cells(cell + ':' + cell1)
objective_value = (Superstructure.MTAC + Superstructure.GlycMTAC + Superstructure.HotUCost + Superstructure.ColdUCost + Superstructure.FeedCost) / 1000
sheet[cell] = 'Objective function: %s' % objective_value
sheet[cell].fill = blue

sheet[get_column_letter(start_col) + str(start_row + h_spacer + 1)] = 'Biodiesel'
sheet[get_column_letter(start_col) + str(start_row + h_spacer + 2)] = 'AIC'
sheet[get_column_letter(start_col) + str(start_row + h_spacer + 3)] = 'TUC'
sheet[get_column_letter(start_col) + str(start_row + h_spacer + 4)] = 'RMC'
sheet[get_column_letter(start_col) + str(start_row + h_spacer + 5)] = 'OMC'
sheet[get_column_letter(start_col) + str(start_row + h_spacer + 6)] = 'WC'
sheet[get_column_letter(start_col) + str(start_row + h_spacer + 7)] = 'TAR'
sheet[get_column_letter(start_col) + str(start_row + h_spacer + 8)] = 'MTAC'

sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 1)] = '1000[$]'
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 2)] = Superstructure.AIC / 1000
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 3)] = Superstructure.TUC / 1000
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 4)] = Superstructure.RMC / 1000
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 5)] = Superstructure.OMC / 1000
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 6)] = Superstructure.WC / 1000
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 7)] = -Superstructure.BDP / 1000
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 8)] = Superstructure.MTAC / 1000


sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 1)] = 'Glycerol'
sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 2)] = 'AIC'
sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 3)] = 'TUC'
sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 4)] = 'RMC'
sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 5)] = 'OMC'
sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 6)] = 'WC'
sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 7)] = 'TAR'
sheet[get_column_letter(start_col+2) + str(start_row + h_spacer + 8)] = 'MTAC'

sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 1)] = '1000[$]'
sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 2)] = Superstructure.GlycAIC / 1000
sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 3)] = Superstructure.GlycTUC/ 1000
sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 4)] = Superstructure.GlycRMC/ 1000
sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 5)] = Superstructure.GlycOMC/ 1000
sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 6)] = Superstructure.GlycWC/ 1000
sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 7)] = -Superstructure.GlycTAR/ 1000
sheet[get_column_letter(start_col+3) + str(start_row + h_spacer + 8)] = Superstructure.GlycMTAC/ 1000


sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 1)] = 'Total'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 2)] = 'AIC'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 3)] = 'TUC'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 4)] = 'RMC'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 5)] = 'OMC'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 6)] = 'WC'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 7)] = 'TAR'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 8)] = 'MTAC'

sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 1)] = '1000[$]'
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 2)] = (Superstructure.GlycAIC + Superstructure.AIC) / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 3)] = (Superstructure.GlycTUC + Superstructure.TUC) / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 4)] = (Superstructure.GlycRMC + Superstructure.RMC) / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 5)] = (Superstructure.GlycOMC + Superstructure.OMC) / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 6)] = (Superstructure.GlycWC + Superstructure.WC) / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 7)] = -(Superstructure.GlycTAR + Superstructure.BDP) / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 8)] = (Superstructure.GlycMTAC + Superstructure.MTAC ) / 1000


sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 10)] = 'General costs'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 11)] = 'Feed'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 12)] = 'HotU'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 13)] = 'ColdU'

sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 15)] = 'Additional info (already taken into account)'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 16)] = 'Membrane Reac'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 17)] = 'ElectricityC'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 18)] = 'HeatingC'
sheet[get_column_letter(start_col+4) + str(start_row + h_spacer + 19)] = 'CoolingC'


sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 10)] = 'General costs'
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 11)] = Superstructure.FeedCost / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 12)] = Superstructure.HotUCost / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 13)] = Superstructure.ColdUCost / 1000

sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 16)] = Superstructure.MembraneReactorCost / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 17)] = Superstructure.ElectricityAmount / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 18)] = Superstructure.HeatingAmount / 1000
sheet[get_column_letter(start_col+5) + str(start_row + h_spacer + 19)] = Superstructure.CoolingAmount / 1000


for a in Superstructure.a:
    for j in Superstructure.j:
        for k in Superstructure.k:
            selected = Superstructure.y.loc[0, (a,j,k)]
            name = Superstructure.equipment_names.loc[k, (a,j)]
            CostCorr = Superstructure.CostCorr.loc[0, (a,j,k)]
            ref_cost = Superstructure.RefCost_data[a,j,k]
            real_cost = CostCorr * ref_cost
            Sizing_data = Superstructure.Sizing_data[a,j,k]
            Ref_index = Superstructure.RefIndex_data[a,j,k]
            
            #Define the 4 corner cells to draw a box in excel
            cell = get_column_letter(col[a - 1]) + str(start_row)
            cell1 = get_column_letter(col[a - 1] + h_size) + str(start_row)
            cell2 = get_column_letter(col[a - 1]) + str(start_row + v_size)
            cell3 = get_column_letter(col[a - 1] + h_size) + str(start_row + v_size)

            #Merge and center the equipment name cell
            if selected >= 0.5:
                sheet.merge_cells(cell + ':' + cell1)
                sheet[cell].alignment = Alignment(horizontal = 'center')
                sheet[cell] = name + str( (a,j,k) )
                
                sheet[get_column_letter(col[a-1]) + str(start_row+1)] = 'Species i'
                sheet[get_column_letter(col[a-1]+1) + str(start_row+1)] = 'Flow [kg/h]'
                sheet[get_column_letter(col[a-1]) + str(start_row+3+len(Superstructure.i))] = 'Total'
                sheet[get_column_letter(col[a-1]+1) + str(start_row+3+len(Superstructure.i))] = Superstructure.Flow_intot.loc[0,(a,j,k)]
                
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 1)] = 'Costs'
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 2)] = 'Ref cost'
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 3)] = 'Ref size'
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 4)] = 'E factor'
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 5)] = 'Ref IDX'
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 6)] = 'Cost corr'
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 7)] = 'Final cost'
                sheet[get_column_letter(col[a-1]+2) + str(start_row + 8)] = 'AIC'
                
            
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 1)] = '1000[$]'
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 2)] = ref_cost / 1000
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 3)] = Superstructure.RefSize_data[a,j,k]
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 4)] = Sizing_data
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 5)] = Ref_index
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 6)] = CostCorr
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 7)] = real_cost / 1000
                sheet[get_column_letter(col[a-1]+3) + str(start_row + 8)] = real_cost * Superstructure.IR_LF / 1000
            
                
                
                for i in Superstructure.i:
                    sheet[get_column_letter(col[a-1]) + str(start_row + i + 1)] = Superstructure.component_names.loc[i]
                    sheet[get_column_letter(col[a-1] + 1) + str(start_row + i + 1)] = Superstructure.Flow_in.loc[i, (a,j,k)]
                    
                sheet[cell].fill = green
                
                for r in range(v_size):
                    for c in range(h_size + 1):
                        cell = get_column_letter(col[a-1] + c) + str(start_row + r + 1)
                        sheet[cell].fill = grey
                
                



workbook.save(filename = filename)




#excel_write(Superstructure,output_file)

